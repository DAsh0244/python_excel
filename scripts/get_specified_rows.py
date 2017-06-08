import os
import sys
import openpyxl
from time import time


def print_row(row_num, row_data, sheet):
    for col,cell in enumerate(row_data, start=1):
        sheet.cell(row=row_num, column=col).value = cell.value


def sec2time(sec, n_msec=3):
    ''' Convert seconds to 'D days, HH:MM:SS.FFF' '''
    if hasattr(sec,'__len__'):
        return [sec2time(s) for s in sec]
    m, s = divmod(sec, 60)
    h, m = divmod(m, 60)
    d, h = divmod(h, 24)
    if n_msec > 0:
        pattern = '%%02d:%%02d:%%0%d.%df' % (n_msec+3, n_msec)
    else:
        pattern = r'%02d:%02d:%02d'
    if d == 0:
        return pattern % (h, m, s)
    return ('%d days, ' + pattern).format(d, h, m, s)

# FILE = 'test.xlsx'
FILE = 'DataDumpReportCopy.xlsx'
COL_NAME_1 = 'A'
TARGET_IDS = 'new_ids.txt'
RESULTS_FILE = 'new_entries.xlsx'
if len(sys.argv) > 1:
    try:
        FILE = sys.argv[1]
        COL_NAME_1 = sys.argv[2]
        DELIM_CHAR = sys.argv[3]
    except:
        pass

t1 = time()

print('getting id keys')
ids = []
with open(TARGET_IDS, 'r') as f:
    for line in f:
        ids.append(int(line))

t2 = time()
print('Operation took {}'.format(sec2time(t2 - t1)))

print('loading workbook {}'.format(FILE))
wb = openpyxl.load_workbook(FILE)
# wb = openpyxl.load_workbook(FILE, read_only=True)
sheets = wb.get_sheet_names()
worksheet = wb.get_sheet_by_name(sheets[0])
title_row = worksheet[1]
col_1 = worksheet[COL_NAME_1]
# col_1 = []
# for row in worksheet.rows:
#    col_1.append(row[0])
new_rows = []

t3 = time()
print('Operation took {}'.format(sec2time(t3 - t2)))

print('finding new entries')
for entry_1 in col_1:
    if entry_1.value in ids:
        new_rows.append(worksheet[entry_1.row])
print('found {} new entries'.format(len(new_rows)))
print('closing workbook')
wb.close()
del wb

t4 = time()
print('Operation took {}'.format(sec2time(t4 - t3)))

print('creating new workbook {}'.format(RESULTS_FILE) )
new_wb = openpyxl.Workbook()
sheet = new_wb['Sheet']
print('writing contents')
print_row(1, title_row, sheet)
for index, row in enumerate(new_rows,start=2):
        print_row(index, row, sheet)
        
print('saving workbook {}'.format(RESULTS_FILE))
new_wb.save(RESULTS_FILE)
new_wb.close()

t5 = time()
print('Operation took {}'.format(sec2time(t5 - t4)))

print('Done!')
print('Total Operation took {}'.format(sec2time(t5-t1)))
sys.exit(0)
