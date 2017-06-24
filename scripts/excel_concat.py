#! /usr/bin/env python
# vim:fileencoding=utf-8
# -*- coding: utf-8 -*-
"""
excel utils
excel_concat.py
Author: Danyal Ahsanullah
Date: 2017, June 16
Copyright (c):  2017 Danyal Ahsanullah
License: MIT (see LICENSE.txt)
Description: CLI tool used to join/split excel files.
"""


import re
import os
import sys
import glob
import logging
import openpyxl
import argparse
import itertools

import common
from common import time_execution

__author__ = 'danyal.ahsanullah@gmail.com'
__version__ = '0.0.2'

# logger creation
formatter = logging.Formatter('%(asctime)s - %(message)s')
logger = logging.getLogger(__file__)
# logger.setLevel(logging.WARNING)  # defaults to logging.WARNING
ch = logging.StreamHandler()
ch.setLevel(logging.WARNING)
ch.setFormatter(formatter)
logger.addHandler(ch)

# cli parser creation
parser = argparse.ArgumentParser('excel_concat')
parser.add_argument('file', type=str, nargs='+', help='file(s) to be read in')
parser.add_argument('--path', '-p', action="store_true",
                    help='if enabled, treats <file> arg as set of paths to dirs of files')
parser.add_argument('--chunk', '-c', type=int, default=1000,
                    help='How many entries to write to new file(s). A size less than zero will write to a single file')
parser.add_argument('--output', '-o', type=str,
                    help='base file name to save as. Resulting files will be incremented on the end.'
                         'If the path given doesnt exist, it will be made')
parser.add_argument('--H_concat', '-H', action="store_true", help='Join files in a sytle of horizontal row appending')
parser.add_argument('--keep_title', '-k', action="store_true",
                    help='Whether to keep title row in each chunked file or not')
parser.add_argument('--order', '-O', type=argparse.FileType('r'),
                    help='Path to file that will dictate file order to be processed in')
parser.add_argument('--no_overwrite', '-n', action="store_true", help='If set, will not overwrite files')
loudness = parser.add_mutually_exclusive_group()

loudness.add_argument('--verbose', '-v', action="count", help='prints a more detailed output')
loudness.add_argument('--quiet', '-q', action="count", help='suppresses console output')
parser.add_argument('--version', action='version', version=__version__)
# parser.add_argument('_args', nargs=argparse.REMAINDER)
# parser.add_argument('file', type=argparse.FileType('r'), nargs='+', help='file(s) name to be read in.')
# parser.add_argument('--sheet', '-s', type=str, nargs='+', help='sheet name to pull data from')


@time_execution(alert='loading workbooks', desc='loading workbooks')
def load_wbs(files, _read_only=True):
    wbs = []
    wss = []
    for file in files:
        book = openpyxl.load_workbook(file, read_only=_read_only)
        ws = book.worksheets[0] 
        common.check_ws_dim(ws)
        wbs.append(book)
        # wss.append(book.worksheets[0])
        wss.append(ws)
        logger.warning('loaded {}\n'.format(file))
    return wbs, wss
    

@time_execution(alert='writing chunks', desc='wrote workbooks')
def write_chunked_wbs(wss, chunk, concat=False, output=None, keep_title=True):
    chunk_num = 0
    if output is None:
        name = 'chunked_workbook'
        ext = 'xlsx'
    else:
        name, ext = output.split('.')
    output = '{}_{:02d}.{}'.format(name, chunk_num, ext)
    if concat:
        for ws1, ws2 in common.pairwise(wss):
            new_wb = openpyxl.Workbook(write_only=True)
            try:
                rows1, rows2 = ws1.rows, ws2.rows
                if keep_title:
                    title_1 = common.row_vals(next(rows1))
                    title_2 = common.row_vals(next(rows2))[2:]  # first two cols are duplicate of cols of first
                    # title_2 = list(filter_list(row_vals(next(rows2)), title_1))
                    title_row = common.flatten([title_1, title_2])
                    # print(title_row)
                while rows1 or rows2:
                    new_wb = openpyxl.Workbook(write_only=True)
                    new_ws = new_wb.create_sheet()
                    if keep_title:
                        new_ws.append(title_row)
                    # try:
                    #     new_ws.append(title_row)
                    # except NameError:
                    #     pass
                    if chunk < 0:
                        while True:
                            # first two cols are duplicate of first two cols of first file
                            new_ws.append(common.flatten([common.row_vals(next(rows1)),
                                                          common.row_vals(next(rows2)[2:])]))
                            # new_ws.append(flatten([row_vals(next(rows1)), row_vals(next(rows2))]))
                    else:
                        for line in range(chunk):
                            # first two cols are duplicate of first two cols of first file
                            new_ws.append(common.flatten([common.row_vals(next(rows1)),
                                                          common.row_vals(next(rows2)[2:])]))
                            # new_ws.append(flatten([row_vals(next(rows1)), row_vals(next(rows2))]))
                    new_wb.save(output)
                    logger.debug('wrote {}'.format(output))
                    chunk_num += 1
                    output = '{}_{:02d}.{}'.format(name, chunk_num, ext)
            except StopIteration:
                new_wb.save(output)
                logger.debug('wrote {}'.format(output))
                chunk_num += 1
                output = '{}_{:02d}.{}'.format(name, chunk_num, ext)
    else:
        if chunk < 0:
            new_wb = openpyxl.Workbook(write_only=True)
            new_ws = new_wb.create_sheet()
            title_row = common.row_vals(next(wss[0].rows))
            new_ws.append(title_row)
            for ws in wss:
                for row in ws.rows:
                    tmp = common.row_vals(row)
                    if tmp == title_row:
                        # next(ws.rows)
                        tmp = common.row_vals(next(ws.rows))
                    new_ws.append(tmp)
            new_wb.save(output)
            logger.debug('wrote {}'.format(output))
        else:
            for ws in wss:
                new_wb = openpyxl.Workbook(write_only=True)
                try:
                    rows1 = ws.rows
                    if keep_title:
                        title_row = common.row_vals(next(rows1))
                        # print(title_row)
                    while rows1:
                        new_wb = openpyxl.Workbook(write_only=True)
                        new_ws = new_wb.create_sheet()
                        try:
                            new_ws.append(title_row)
                        except NameError:
                            pass
                        for line in range(chunk):
                            new_ws.append(common.row_vals(next(rows1)))
                        new_wb.save(output)
                        logger.debug('wrote {}'.format(output))
                        chunk_num += 1
                        output = '{}_{:02d}.{}'.format(name, chunk_num, ext)
                except StopIteration:
                    new_wb.save(output)
                    logger.debug('wrote {}'.format(output))
                    chunk_num += 1
                    output = '{}_{:02d}.{}'.format(name, chunk_num, ext)

@time_execution('Total operation')
def main():
    args = parser.parse_args()

    if args.verbose:
        logger.setLevel(logging.DEBUG)
        ch.setLevel(logging.DEBUG)
        logger.warning('Operating in verbose mode:')
    elif args.quiet:
        logger.setLevel(logging.NOTSET)
        ch.setLevel(logging.NOTSET)
    logger.debug(vars(args))

    if args.path:
        files = []
        for entry in args.file:
            files.append(glob.glob(entry))
        args.file = list(common.filter_list(common.flatten(files)))
        logger.debug('The following files were found: {}'.format(args.file))

    if args.order:
        try: 
            order = args.order.read().splitlines()
            args.file = common.sorted_files(args.file, order)
            logger.debug('files successfully reordered')
        except IOError:
            raise IOError('File:{} could not be read'.format(args.output.name))

    for file in iter(args.file):
        if not os.path.exists(file):
            raise OSError('file not found: {}'.format(file))
    logger.debug('files were all found')

    if args.output:
        os.makedirs(os.path.dirname(args.output), exist_ok=True)
        logger.debug('Output path set')

    workbooks, worksheets = load_wbs(args.file)

    write_chunked_wbs(worksheets, args.chunk, concat=args.H_concat, 
                      output=args.output, keep_title=args.keep_title)

                      
if __name__ == '__main__':
    main()
    sys.exit(0)
