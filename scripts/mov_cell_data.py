import os
import sys
import openpyxl
import argparse
import re
import time

from common import *


__version__ = '0.0.1'

# cli parser definition
parser = argparse.ArgumentParser()
parser.add_argument('--version', action='version', version=__version__)
parser.add_argument('file', type=str, help='file name to be read in.')
parser.add_argument('dict', type=str, help='file for string replacement dictionary file')
parser.add_argument('col_name', type=str, help='Base column to scan and move entries from')
parser.add_argument('--delim', type=str, help='delimiter character to break on', default=',')
parser.add_argument('--sheet', '-s', type=str, help='sheet name to pull data from')
parser.add_argument('--output', '-o', type=str, help='file name to save as')
parser.add_argument('--direct_col', '-d', action="store_true",  # default=False, 
                    help='use direct column name instead of name defined in first row of column')


@time_execution
def load_move_dict(file_path):
    comment = '###'
    mov_dict = {}
    with open(file_path,'r') as f:
        for line in f:
            if line.startswith(comment) or not line.strip():
                continue
            else:
                ln = line.strip()
                if ln.endswith(':'):
                    column_name = ln[:-1].strip()
                    # column_name = ln
                    sub_dict = {}
                    nxt_line = next(f)
                    while nxt_line.strip() != '':
                        if nxt_line.startswith(comment):
                            continue
                        else:
                            (key,val) = nxt_line.split('\t')
                            sub_dict[key.strip()] = val.strip('\r\n')
                        nxt_line = next(f)
                    if sub_dict:
                        mov_dict[column_name] = sub_dict
    return mov_dict 


def mov_data(target, source_list, destination, delim):
    tmp = source_list
    if target in source_list:
        tmp = list(filter(lambda x:x != target, tmp))
        if destination.value is not None:
            destination.value += '{} {}'.format(delim, target)
        else:
            destination.value = '{}'.format(target)
    return tmp

@time_execution    
def move_col_entries(ws, base_col, delimiter, mov_dict, col_mapping):
    rev_title_mapping = inv_map(col_mapping)
    col_name = base_col[0].column
    for cell in base_col:
        try:
            tmp_list = cell.value.split(delimiter)
            tmp_list = [entry.strip() for entry in tmp_list]            
            if rev_title_mapping[col_name] in mov_dict:
                sub_dict = mov_dict[rev_title_mapping[col_name]] 
                # print('sub_dict:\n', sub_dict)
                for key, value in sub_dict.items():
                    if key in tmp_list:
                        dest = ws[col_mapping[value] + str(cell.row)]
                        tmp_list = mov_data(key, tmp_list, dest, delimiter)
                tmp = set(tmp_list)            
            new_str = '{} '.format(delimiter).join(tmp)
            cell.value = new_str
        except Exception as e:
            continue


@time_execution    
def main():
    args = parser.parse_args()
    mov_dict = load_move_dict(args.dict)
     
    wb, ws = get_workbook(args)
    title_mapping = excel_mappings(ws[1])

    if args.direct_col:
        base_col = ws[args.col_name]
    else:
        base_col = ws[title_mapping[args.col_name]]
    
    move_col_entries(ws, base_col, args.delim, mov_dict, title_mapping)
    
    write_workbook(wb, args)

if __name__ == '__main__':
    main()
    sys.exit(0)
    