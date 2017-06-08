import os
import sys
import openpyxl
import argparse
import re

from common import *

__version__ = '0.0.1'

# cli parser definition
parser = argparse.ArgumentParser()
parser.add_argument('--version', action='version', version=__version__)
parser.add_argument('file', type=str, help='file name to be read in.')
parser.add_argument('dict', type=str, help='file for string replacement dictionary file')
parser.add_argument('col_name', type=str, help='Column to perform string replacement on')
parser.add_argument('--delim', type=str, help='delimiter character to break on', default=',')
parser.add_argument('--sheet', '-s', type=str, help='sheet name to pull data from')
parser.add_argument('--output', '-o', type=str, help='file name to save as')
parser.add_argument('--direct_col', '-d', action="store_true",  # default=False, 
                    help='use direct column name instead of name defined in first row of column')


@time_execution
def load_replace_dict(file_path):
    comment = '###'
    replace_dict = {}
    with open(file_path,'r') as f:
        for line in f:
            if line.startswith(comment) or not line.strip():
                continue
            else:
                (key,val) = line.split('\t')
                replace_dict[key] = val.strip('\r\n')
    return replace_dict


# @time_execution
def mapped_replace(bad_list, replace_dict, join_char=','):
    new_list = [entry.strip() for entry in bad_list]
    new_list = list(filter(lambda x:x != ' ', new_list))
    tmp_str = join_char.join(new_list)
    pattern = re.compile(r'\b(' + 
                        '|'.join(re.escape(key) for key in sorted(replace_dict.keys(), key=len, reverse=True)) +
                        r')\b')
    tmp_str = pattern.sub(lambda x: replace_dict[x.group()], tmp_str)
    new_list = tmp_str.split(join_char)
    # for key in replace_dict:
        # new_list = [w.replace(key, replace_dict[key]) for w in tmp_list]
    new_list = list(filter(None, new_list))
    return new_list


@time_execution
def str_replace(column, delimiter, replace_dict):
    unique = set()
    for cell in column:
        # print(cell.value)
        try:
            tmp_list = cell.value.split(delimiter)
            tmp_list = mapped_replace(tmp_list, replace_dict)
            # print(tmp_list)
            
            # if any(move_list in tmp_list):
            # if 'Impact of Cancer on the Family' in tmp_list:
                # tmp_list = list(filter(lambda x:x != 'Impact of Cancer on the Family', tmp_list))
                # mov_cell.value += ', Impact of Cancer on the Family'
            tmp = set(tmp_list)
            for entry in tmp:
                unique.add(entry)
            new_str = '{} '.format(delimiter).join(tmp)
            cell.value = new_str
        except:
            continue
    return unique
    
    
# main function 
@time_execution
def main():
    args = parser.parse_args()
    replacement_mappings = load_replace_dict(args.dict)
    
    wb, ws = get_workbook(args)
    title_mapping = excel_mappings(ws[1])

    if args.direct_col:
        col = ws[args.col_name]
    else:
        col = ws[title_mapping[args.col_name]]
    unique = str_replace(col, args.delim, replacement_mappings)
        
    # print('found {} unique entries:'.format(len(unique)))
    # for entry in unique:
        # print(entry)
    
    write_workbook(wb, args)

    
if __name__ == '__main__':
    main()
    sys.exit(0)
