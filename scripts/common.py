#! /usr/bin/env python
# vim:fileencoding=utf-8
# -*- coding: utf-8 -*-
"""
excel utils
common.py
Author: Danyal Ahsanullah
Date: 2017, June 16
Copyright (c):  2017 Danyal Ahsanullah
License: MIT (see LICENSE.txt)
Description: common functions that will be used between scripts
"""

import time as _time
import openpyxl as _openpyxl
import string as _string
import itertools as _itertools
import os as _os
import logging as _logging
from openpyxl.utils.cell import *


__author__ = 'danyal.ahsanullah@gmail.com'
__version__ = '0.0.2'


_formatter = _logging.Formatter('%(asctime)s - %(message)s')
# _logger = _logging.getLogger(__file__)
_logger = _logging.getLogger('A')
# logger.setLevel(logging.DEBUG)  # defaults to logging.WARNING
_ch = _logging.StreamHandler()
_ch.setLevel(_logging.DEBUG)
_ch.setFormatter(_formatter)
_logger.addHandler(_ch)
_logger.setLevel(_logging.DEBUG)


def ensure_dir(file_path):
    directory = _os.path.dirname(file_path)
    if not _os.path.exists(directory):
        _os.makedirs(directory)


def path_leaf(path):
    head, tail = _os.path.split(path)
    return tail or _os.path.basename(head)


def sorted_files(files, order):
    # key_func = lambda x: order.index(path_leaf(x))
    def key_func(x):
        return order.index(path_leaf(x))
    if any([path_leaf(entry) not in order for entry in files]):
        raise ValueError('file(s) given not listed in order.')
    if any([entry not in [path_leaf(file) for file in files] for entry in order]):
        raise ValueError('file(s) in order missing in provided files.')
    return sorted(files, key=key_func)


def pairwise(iterable):
    """"s -> (s0,s1), (s1,s2), (s2, s3), ..."""
    a, b = _itertools.tee(iterable)
    next(b, None)
    return zip(a, b)


def grouped(iterable, n):
    """s -> (s0,s1,s2,...sn-1), (sn,sn+1,sn+2,...s2n-1), (s2n,s2n+1,s2n+2,...s3n-1), ..."""
    return zip(*[iter(iterable)] * n)


def filter_list(full_list, excludes=(None, '')):
    s = set(excludes)
    return (x for x in full_list if x not in s)


def flatten(iterable, n=1):
    """
    remove n levels of list nesting for a list of lists
    assumes uniform nesting
    :param iterable: nested list of lists to be flattened
    :param n: how many levels of nesting to remove
    :return: flattened list
    """
    tmp = iterable
    for _ in range(n):
        tmp = [item for sublist in tmp for item in sublist]
    return tmp


def _sec2time(sec, n_msec=5):
    """
    Convert seconds to "D days, HH:MM:SS.FFF"
    :param sec: number of seconds
    :param n_msec: number of milliseconds to use
    :return: string representing of time in human readable format
    """
    if hasattr(sec, '__len__'):
        return [_sec2time(s) for s in sec]
    m, s = divmod(sec, 60)
    h, m = divmod(m, 60)
    d, h = divmod(h, 24)
    d, h, m = int(d), int(h), int(m)
    if n_msec > 0:
        pattern = '{{:02d}}:{{:02d}}:{{:0{space}.{msec}f}}'.format(space=n_msec + 3, msec=n_msec)
        # pattern = '%%02d:%%02d:%%0%d.%df' %(n_msec+3, n_msec)
    else:
        pattern = '{:02d}:{:02d}:{:02d}'
        # pattern = r'%02d:%02d:%02d'
    if d == 0:
        # return pattern % (h, m, s)
        return pattern.format(h, m, s)
    return ('{:02d} days, ' + pattern).format(d, h, m, s)


def time_execution(desc=None, alert=None):
    """
    decorator to measure function execution time
    :param desc: string to be used inplace of function name
    :param alert: string to be displayed before function call
    :return: N/A
    """
    def decorator(method):
        def wrapper(*args, **kw):
            if alert:
                # _logger.info('\n{}:'.format(alert.title()))
                print('\n{}:'.format(alert.title()))
            t1 = _time.perf_counter()
            result = method(*args, **kw)
            t2 = _time.perf_counter()
            time_elapsed = _sec2time(t2 - t1)
            if desc:
                msg = '{} took {}'.format(desc, time_elapsed)
            else:
                msg = '"{}": Operation took {}'.format(method.__name__,  time_elapsed)
            _logger.info(msg)
            return result
        return wrapper
    return decorator


def write_row(row_num, row_data, sheet):
    """
    writes a row of data to the specified row num
    :param row_num: row num to write row data to
    :param row_data: data to be written to the row
    :param sheet: sheet to write row to
    :return: N/A
    """
    for col, cell in enumerate(row_data, start=1):
        sheet.cell(row=row_num, column=col).value = cell.value


def excel_mappings(title_row):
    """
    creates a dictionary that shows mappings between cell contents and cell columns
    :param title_row: list of openpyxl.cell.Cell
    :return: dict of col names to col vals
    """
    return {entry.value.strip(): entry.column for entry in title_row}


def inv_map(dictionary: dict):
    """
    creates a inverse mapped dictionary of the provided dict
    :param dictionary: dictionary to be reverse mapped
    :return: inverse mapped dict
    """
    return {v: k for k, v in dictionary.items()}


def increment_char(c):
    """
    Increment an uppercase character, returning 'A' if 'Z' is given
    """
    return chr(ord(c) + 1) if c != 'Z' else 'A'


def increment_str(s):
    lpart = s.rstrip('Z')
    num_replacements = len(s) - len(lpart)
    new_s = lpart[:-1] + increment_char(lpart[-1]) if lpart else 'A'
    new_s += 'A' * num_replacements
    return new_s


def col2num(col):
    num = 0
    for c in col:
        if c in _string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num


def colnum_string(n):
    div = n
    string = ""
    while div > 0:
        mod = (div - 1) % 26
        string = chr(65 + mod) + string
        div = int((div - mod) / 26)
    return string


def row_vals(row, gen=False):
    """
    return a list of values for a list of openpyxl.cell.Cell
    :param row: list of openpyxl.cell.Cell
    :param gen: if true, returns generator object
    :return: N/A
    """
    if gen:
        return (cell.value for cell in row)
    else:
        return [cell.value for cell in row]


def get_cols_by_name(start, stop, inclusive=True):
    if inclusive:
        cols = list(map(colnum_string, range(col2num(start), col2num(stop) + 1)))
    else:
        cols = list(map(colnum_string, range(col2num(start), col2num(stop))))
    return cols


def format_title_row(row, cols, tag='Interest', tag_label=':'):
    leading_tag = '{}{} '.format(tag, tag_label)
    for cell in row:
        if cell.column in cols and not cell.value.startswith(leading_tag):
            cell.value = leading_tag + cell.value


@time_execution('getting workbook')
def get_workbook(args):
    wb = _openpyxl.load_workbook(args.file)
    if not args.sheet:
        ws = wb.worksheets[0]
    else:
        ws = wb.get_sheet_by_name(args.sheet)
    return wb, ws


@time_execution('writing workbook')
def write_workbook(wb, args):
    if not args.output:
        wb.save(args.file)
    else:
        wb.save(args.output)


# @time_execution
def check_ws_dim(sheet):
    if sheet.calculate_dimension() == 'A1:A1':
        _logger.warning('worksheet dimensions incorrect, recalculating...')
        sheet.max_row = sheet.max_column = None
        dim = sheet.calculate_dimension(force=True)
        _logger.warning('recalculated dimensions as {}'.format(dim))
        _logger.warning('press <Ctrl+c> to cancel if dimensions are incorrect')


''' EOF '''
