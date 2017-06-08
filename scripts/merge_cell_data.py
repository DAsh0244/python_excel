import os
import sys
import openpyxl
import argparse
import re
import time

from common import *


__version__ = '0.0.1'

# cli parser definition
parser = argparse.ArgumentParser()
parser.add_argument('--version', action='version', version=__version__)
parser.add_argument('file', type=str, help='file name to be read in.')
parser.add_argument('dict', type=str, help='file for string replacement dictionary file')
parser.add_argument('--delim', type=str, help='delimiter character to break on', default=',')
parser.add_argument('--sheet', '-s', type=str, help='sheet name to pull data from')
parser.add_argument('--output', '-o', type=str, help='file name to save as')


@time_execution
def load_merge_dict(file_path):
    comment = '###'
    merge_dict = {}
    with open(file_path,'r') as f:
        for line in f:
            if line.startswith(comment) or not line.strip():
                continue
            else:
                ln = line.strip()
                if ln.endswith(':'):
                    column_name = ln[:-1].strip()
                    # column_name = ln
                    merge_cols = []
                    nxt_line = next(f)
                    while nxt_line.strip() != '':
                        if nxt_line.startswith(comment):
                            continue
                        else:
                            merge_cols.append(nxt_line.strip('\r\n'))
                        nxt_line = next(f)
                    if merge_cols:
                        merge_dict[column_name] = merge_cols
    return merge_dict 


def merge_cells_contents(source_cell, merge_map, delim=','):
    merge_cells = []
    for col in merge_map:
        merge_cells.append(ws[col+str(source_cell.row)].value)
    return set(merge_cells)

@time_execution    
def merge_col_entries(ws, col_mapping, merge_dict, delimiter=','):
    rev_col_mapping = inv_map(col_mapping)
    merge_map = {col_mapping[k]: [col_mapping[entry] for entry in v] for k,v in merge_dict.items()}
    print(merge_map)
    
    for key in merge_map:
        for cell in ws[key]:  # selecting cells from col 
            try:
                tmp = merge_cells_contents(cell, merge_map)
                new_str = '{} '.format(delimiter).join(tmp)
                
                cell.value = new_str
            except Exception as e:
                continue


@time_execution    
def main():
    args = parser.parse_args()
    merge_dict = load_merge_dict(args.dict)
     
    wb, ws = get_workbook(args)
    title_mapping = excel_mappings(ws[1])
    rev_col_mapping = inv_map(title_mapping)
    
    merge_col_entries(ws, title_mapping, merge_dict, args.delim)
    
    write_workbook(wb, args)

if __name__ == '__main__':
    main()
    sys.exit(0)
    